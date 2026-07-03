import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# ── NUEVO HORARIO (Ana franco Dom+Lun) ──────────────────────────────────────
# dow: 0=Lun,1=Mar,2=Mié,3=Jue,4=Vie,5=Sáb,6=Dom
SCHEDULE = {
    "Ana":      {0:None,   1:"13-22",2:"12-21",3:"12-22",4:"12-22",5:"12-22",6:None},
    "Franco":   {0:"13-22",1:None,   2:None,   3:"12-22",4:"9-19", 5:"9-19", 6:"13-22"},
    "Yazmin":   {0:"12-21",1:"12-21",2:None,   3:None,   4:"9-19", 5:"9-19", 6:"12-22"},
    "Martina":  {0:"9-17", 1:"9-17", 2:"9-17", 3:"9-17", 4:None,   5:None,   6:None},
    "Victoria": {0:"14-22",1:"14-22",2:"14-22",3:None,   4:None,   5:None,   6:"10-18"},
}

COLORS  = {"Ana":"9FC5E8","Franco":"EA9999","Yazmin":"93C47D","Martina":"F6B26B","Victoria":"CECBF6"}
COLORST = {"Ana":"0C447C","Franco":"712B13","Yazmin":"27500A","Martina":"633806","Victoria":"3C3489"}

FERIADOS  = {datetime.date(2026,7,9): "Feriado Nacional"}
DAY_ES    = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
EMP_ORDER = ["Ana","Franco","Yazmin","Martina","Victoria"]
TIMES     = [(h, f"{h:02d}:00-{h+1:02d}:00") for h in range(9,22)]  # 09→22

# ── HELPERS ─────────────────────────────────────────────────────────────────
def fl(c): return PatternFill("solid", start_color=c, end_color=c)
def fn(bold=False, sz=9, col="000000"): return Font(name="Arial", bold=bold, size=sz, color=col)
def al(h="center",v="center"): return Alignment(horizontal=h, vertical=v, wrap_text=False)

T   = Side(style="thin",   color="DEDEDE")
TM  = Side(style="medium", color="A0A0A0")
TW  = Side(style="medium", color="FFFFFF")

def shift_hrs(sh):
    if not sh: return set()
    s,e = map(int,sh.split("-")); return set(range(s,e))

def workers_at(dow, hour):
    w = [(p, SCHEDULE[p][dow]) for p in EMP_ORDER if SCHEDULE[p].get(dow) and hour in shift_hrs(SCHEDULE[p][dow])]
    return sorted(w, key=lambda x: int(x[1].split("-")[0]))

# Máximo de personas simultáneas (para fijar sub-filas por franja)
MAX_W = max(len(workers_at(d,h)) for h,_ in TIMES for d in range(7))
# MAX_W = 4

# ── CREAR HOJA SEMANAL ───────────────────────────────────────────────────────
def make_sheet(wb, week_days, sname):
    ws = wb.create_sheet(sname)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale     = 90

    N_DAYS = len(week_days)
    HDR    = 2          # filas de encabezado
    SUB    = MAX_W      # sub-filas por franja horaria

    # Anchos de columna
    ws.column_dimensions["A"].width = 13.5
    for i in range(N_DAYS):
        ws.column_dimensions[get_column_letter(i+2)].width = 13.5

    # Alturas de fila
    ws.row_dimensions[1].height = 24  # título
    ws.row_dimensions[2].height = 20  # días
    for r in range(3, 3 + len(TIMES)*SUB + 1):
        ws.row_dimensions[r].height = 14

    # ── FILA 1: título ──
    last_col = get_column_letter(N_DAYS + 1)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = "GUARDIA ADMINISTRATIVA — JULIO 2026"
    c.font      = fn(bold=True, sz=12, col="FFFFFF")
    c.fill      = fl("4A4F5C")
    c.alignment = al()

    # ── FILA 2: encabezados de día ──
    c = ws.cell(row=2, column=1, value="Horario")
    c.font = fn(bold=True, sz=9, col="FFFFFF"); c.fill = fl("4A4F5C"); c.alignment = al()

    for col_i, (date, dow) in enumerate(week_days):
        col_num = col_i + 2
        is_fe   = date in FERIADOS
        is_wk   = dow >= 5
        bg      = "C0392B" if is_fe else ("504E86" if is_wk else "3ECDB0")
        label   = f"{DAY_ES[dow]}  {date.strftime('%-d/%-m')}"
        if is_fe: label += "  ⚑"
        c = ws.cell(row=2, column=col_num, value=label)
        c.font      = fn(bold=True, sz=10, col="FFFFFF")
        c.fill      = fl(bg)
        c.alignment = al()
        c.border    = Border(left=TW, right=TW, bottom=TW)

    # ── FRANJAS HORARIAS ──
    for t_idx, (hour, slot) in enumerate(TIMES):
        r0     = HDR + 1 + t_idx * SUB          # primera sub-fila de esta franja
        r_last = r0 + SUB - 1                    # última sub-fila

        is_last_slot = (t_idx == len(TIMES) - 1)
        bot_heavy    = Side(style="medium", color="AAAAAA")
        bot_light    = Side(style="thin",   color="DEDEDE")
        slot_bot     = bot_light if is_last_slot else bot_heavy

        # ── Celda de horario (columna A, merged verticalmente) ──
        if r0 < r_last:
            ws.merge_cells(f"A{r0}:A{r_last}")
        c = ws.cell(row=r0, column=1, value=slot)
        c.font      = fn(sz=9, col="4A4F5C", bold=True)
        c.fill      = fl("EBF5FB")
        c.alignment = al()
        c.border    = Border(
            top=bot_heavy if t_idx > 0 else Side(style="medium", color="888888"),
            bottom=slot_bot,
            left=Side(style="medium", color="888888"),
            right=Side(style="medium", color="AAAAAA"),
        )

        # ── Celdas de colaboradores por día ──
        for col_i, (date, dow) in enumerate(week_days):
            col_num = col_i + 2
            active  = workers_at(dow, hour)

            for sub in range(SUB):
                row = r0 + sub
                c   = ws.cell(row=row, column=col_num)

                is_last_sub = (sub == SUB - 1)
                is_first    = (col_i == 0)

                c.border = Border(
                    top=bot_heavy if (t_idx > 0 and sub == 0) else bot_light,
                    bottom=slot_bot if is_last_sub else bot_light,
                    left=Side(style="medium", color="AAAAAA") if is_first else bot_light,
                    right=Side(style="medium", color="AAAAAA") if col_i == N_DAYS-1 else bot_light,
                )

                if sub < len(active):
                    person, shift = active[sub]
                    c.value     = person
                    c.font      = fn(bold=True, sz=9, col=COLORST[person])
                    c.fill      = fl(COLORS[person])
                    c.alignment = al()
                else:
                    c.value = ""
                    c.fill  = fl("FAFAFA" if sub % 2 == 0 else "F5F5F5")

    ws.freeze_panes = "B3"
    return ws

# ── GENERAR SEMANAS ──────────────────────────────────────────────────────────
def make_wk(start, n):
    return [(start + datetime.timedelta(days=i),
             (start + datetime.timedelta(days=i)).weekday())
            for i in range(n)]

WEEKS = [
    ("01-05 Jul", make_wk(datetime.date(2026,7,1),  5)),
    ("06-12 Jul", make_wk(datetime.date(2026,7,6),  7)),
    ("13-19 Jul", make_wk(datetime.date(2026,7,13), 7)),
    ("20-26 Jul", make_wk(datetime.date(2026,7,20), 7)),
    ("27-31 Jul", make_wk(datetime.date(2026,7,27), 5)),
]

wb = openpyxl.Workbook()
wb.remove(wb.active)
for name, week in WEEKS:
    make_sheet(wb, week, name)

out = "/mnt/user-data/outputs/guardia_julio_2026_cuadro.xlsx"
wb.save(out)
print(f"OK → {out}  |  MAX simultáneos: {MAX_W} personas")
