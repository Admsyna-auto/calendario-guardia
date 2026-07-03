import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# ── CONFIG ────────────────────────────────────────────────────────────────────
COLORS = {
    "Ana":      ("9FC5E8","0C447C"),
    "Franco":   ("EA9999","712B13"),
    "Yazmin":   ("93C47D","27500A"),
    "Martina":  ("F6B26B","633806"),
    "Victoria": ("CECBF6","3C3489"),
}

schedule = {
    "Ana":      {0:None,   1:None,   2:"12-21",3:"13-22",4:"12-22",5:"12-22",6:"10-20"},
    "Franco":   {0:"13-22",1:None,   2:None,   3:"9-19", 4:"9-19", 5:"9-19", 6:"13-22"},
    "Yazmin":   {0:"12-21",1:"12-21",2:None,   3:None,   4:"9-19", 5:"9-19", 6:"12-22"},
    "Martina":  {0:"9-17", 1:"9-17", 2:"9-17", 3:"9-17", 4:None,   5:None,   6:None},
    "Victoria": {0:"14-22",1:"14-22",2:"14-22",3:"14-22",4:None,   5:None,   6:None},
}

FERIADOS  = {datetime.date(2026, 7, 9)}
DAY_ES    = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
EMP_ORDER = ["Ana","Franco","Yazmin","Martina","Victoria"]
TIMES     = [(h, f"{h:02d}:00-{h+1:02d}:00") for h in range(9, 22)]   # 9→22

# ── HELPERS ───────────────────────────────────────────────────────────────────
def fl(c):  return PatternFill("solid", start_color=c, end_color=c)
def fn(bold=False, sz=9, col="000000"): return Font(name="Arial", bold=bold, size=sz, color=col)
def al(h="center"): return Alignment(horizontal=h, vertical="center")

thin   = Side(style="thin",   color="D8D8D8")
thin_r = Side(style="thin",   color="AAAAAA")
CELL_B = Border(top=thin, bottom=thin, left=thin, right=thin)

def shift_hours(sh):
    if not sh: return set()
    s, e = map(int, sh.split("-")); return set(range(s, e))

def workers_on(dow):
    w = [(p, schedule[p][dow]) for p in EMP_ORDER if schedule[p].get(dow)]
    return sorted(w, key=lambda x: int(x[1].split("-")[0]))

# ── BUILD ONE WEEK SHEET ──────────────────────────────────────────────────────
def make_sheet(wb, week_days, sheet_name):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    # ── Build column map ──
    day_meta = []   # (date, dow, first_col, workers)
    col = 1         # col 1 = time label

    for date, dow in week_days:
        workers = workers_on(dow)
        if not workers:
            continue
        p_start = col + 1
        day_meta.append((date, dow, p_start, workers))
        col = p_start + len(workers) - 1

    # Column widths
    ws.column_dimensions["A"].width = 12.5
    for _, _, p_start, workers in day_meta:
        for i in range(len(workers)):
            ws.column_dimensions[get_column_letter(p_start + i)].width = 9.2

    # Row heights
    ws.row_dimensions[1].height = 22
    for r in range(2, 2 + len(TIMES)):
        ws.row_dimensions[r].height = 16

    # ── ROW 1: day headers ──
    c = ws.cell(row=1, column=1, value="Horario")
    c.font = fn(bold=True, sz=9, col="FFFFFF")
    c.fill = fl("4A4F5C")
    c.alignment = al()

    for date, dow, p_start, workers in day_meta:
        last = p_start + len(workers) - 1
        is_fe = date in FERIADOS
        is_wk = dow >= 5
        hdr_bg = "C0392B" if is_fe else ("504E86" if is_wk else "3ECDB0")
        label  = f"{DAY_ES[dow]}  {date.strftime('%-d/%-m')}"
        if is_fe: label += "  ⚑"
        if p_start < last:
            ws.merge_cells(f"{get_column_letter(p_start)}1:{get_column_letter(last)}1")
        c = ws.cell(row=1, column=p_start, value=label)
        c.font = fn(bold=True, sz=10, col="FFFFFF")
        c.fill = fl(hdr_bg)
        c.alignment = al()

    # ── ROWS 2+: time slots ──
    for r_idx, (hour, slot) in enumerate(TIMES):
        row = 2 + r_idx

        # Time label
        c = ws.cell(row=row, column=1, value=slot)
        c.font = fn(sz=9, col="4A4F5C")
        c.fill = fl("EBF5FB")
        c.alignment = al()
        c.border = Border(right=thin_r, bottom=thin)

        # Person cells
        for date, dow, p_start, workers in day_meta:
            for i, (person, shift) in enumerate(workers):
                col_n = p_start + i
                c     = ws.cell(row=row, column=col_n)
                bg_c, txt_c = COLORS[person]

                if hour in shift_hours(shift):
                    c.value = person
                    c.font  = fn(bold=True, sz=9, col=txt_c)
                    c.fill  = fl(bg_c)
                else:
                    c.value = ""
                    c.fill  = fl("F7F7F7")
                c.alignment = al()
                c.border    = CELL_B

    # Freeze panes: keep time col + header row visible
    ws.freeze_panes = "B2"
    return ws

# ── DEFINE WEEKS ──────────────────────────────────────────────────────────────
def make_week(start_date, ndays):
    return [(start_date + datetime.timedelta(days=i),
             (start_date + datetime.timedelta(days=i)).weekday())
            for i in range(ndays)]

WEEKS = [
    ("01-05 Jul", make_week(datetime.date(2026, 7,  1), 5)),
    ("06-12 Jul", make_week(datetime.date(2026, 7,  6), 7)),
    ("13-19 Jul", make_week(datetime.date(2026, 7, 13), 7)),
    ("20-26 Jul", make_week(datetime.date(2026, 7, 20), 7)),
    ("27-31 Jul", make_week(datetime.date(2026, 7, 27), 5)),
]

wb = openpyxl.Workbook()
wb.remove(wb.active)
for name, week in WEEKS:
    make_sheet(wb, week, name)

out = "/mnt/user-data/outputs/guardia_julio_2026_franja.xlsx"
wb.save(out)
print("OK →", out)
