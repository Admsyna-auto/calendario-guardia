import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

wb = openpyxl.Workbook()

# ── CONFIG ────────────────────────────────────────────────────────────────────
# Plantilla semanal (dow: 0=Lun … 6=Dom). None = día franco.
schedule = {
    "Ana":      {0: None,    1: None,    2: "13-22", 3: "13-22", 4: "12-21", 5: "12-21", 6: "13-22"},
    "Franco":   {0: "13-22", 1: "13-22", 2: None,    3: None,    4: "9-19",  5: "9-19",  6: "12-22"},
    "Yazmin":   {0: "12-21", 1: "12-21", 2: "12-21", 3: "9-18",  4: "9-18",  5: None,    6: None},
    "Martina":  {0: "9-17",  1: "9-17",  2: "9-17",  3: "9-17",  4: None,    5: None,    6: None},
    "Victoria": {0: None,    1: None,    2: None,     3: "14-22", 4: "14-22", 5: "14-22", 6: "10-18"},
}

FERIADOS = {datetime.date(2026, 7, 9): "Día de la Independencia"}
EMPLOYEES = ["Ana", "Franco", "Yazmin", "Martina", "Victoria"]
DAYS_ES   = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
START     = datetime.date(2026, 7, 1)

HRS = {"9-17": 8, "9-18": 9, "9-19": 10,
       "10-18": 8, "12-21": 9, "12-22": 10,
       "13-22": 9, "14-22": 8}

CONTRACT  = {"Ana": 45, "Franco": 48, "Yazmin": 45, "Martina": 32, "Victoria": 32}
FRANCOS   = {"Ana": "Lun + Mar (fijo)", "Franco": "Mié + Jue",
             "Yazmin": "Sáb + Dom", "Martina": "Vie + Sáb + Dom",
             "Victoria": "Lun + Mar + Mié"}
TURNOS    = {"Ana":     "Centro/Tarde (12-21 / 13-22)",
             "Franco":  "Tarde (13-22) · Pico (9-19) · Dom (12-22)",
             "Yazmin":  "Centro (12-21) · Mañana (9-18)",
             "Martina": "Mañana (9-17)",
             "Victoria":"Noche (14-22) · Dom (10-18)"}

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
def fl(c): return PatternFill("solid", start_color=c, end_color=c)
def fn(bold=False, sz=10, col="000000", it=False):
    return Font(name="Arial", bold=bold, size=sz, color=col, italic=it)
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

C_HDR   = "1F3864"; C_HDR2 = "2F5496"
C_OFF   = "D9D9D9"; C_FER  = "FFF2CC"
C_WKD   = "F5F5F5"; C_PEAK = "FFFBE6"
C_MAN   = "C6EFCE"; C_TC   = "BDD7EE"
C_TARD  = "FFEB9C"; C_NOCH = "FCE4D6"
C_TIT   = "EBF5FB"; C_WHI  = "FFFFFF"

def shift_fill(shift):
    if not shift: return fl(C_OFF)
    if shift[0] in "9" or shift[:2] == "10": return fl(C_MAN)
    if shift[:2] == "12": return fl(C_TC)
    if shift[:2] == "13": return fl(C_TARD)
    if shift[:2] == "14": return fl(C_NOCH)
    return fl(C_WHI)

def hdr_cell(ws, row, col, val, bg=C_HDR, sz=9, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = fn(bold=True, sz=sz, col="FFFFFF")
    c.fill = fl(bg)
    c.alignment = al(wrap=wrap)
    c.border = bdr()
    return c

def data_cell(ws, row, col, val, bg=C_WHI, bold=False, sz=9, col_font="000000", h="center", it=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = fn(bold=bold, sz=sz, col=col_font, it=it)
    c.fill = fl(bg)
    c.alignment = al(h=h)
    c.border = bdr()
    return c

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — CALENDARIO MENSUAL
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Julio 2026"
ws1.freeze_panes = "C5"
ws1.sheet_view.showGridLines = False

for i, w in enumerate([10,6,13,13,13,13,13,9,23], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title row
ws1.row_dimensions[1].height = 32
ws1.merge_cells("A1:I1")
c = ws1["A1"]
c.value = "GUARDIA ADMINISTRATIVA  ·  JULIO 2026"
c.font = fn(bold=True, sz=15, col=C_HDR)
c.fill = fl(C_TIT)
c.alignment = al()

ws1.row_dimensions[2].height = 15
ws1.merge_cells("A2:I2")
c = ws1["A2"]
c.value = "Riiing / Syna S.A.  —  Distribución de turnos y francos"
c.font = fn(it=True, sz=9, col="595959")
c.alignment = al()

ws1.row_dimensions[3].height = 5

# Header row
ws1.row_dimensions[4].height = 32
hdrs = ["Fecha","Día","Ana\n(45 hs)","Franco\n(48 hs)","Yazmin\n(45 hs)","Martina\n(32 hs)","Victoria\n(32 hs)","Activos","Observaciones"]
for col, h in enumerate(hdrs, 1):
    hdr_cell(ws1, 4, col, h, wrap=True)

# Data rows
for d in range(31):
    date  = START + datetime.timedelta(days=d)
    row   = 5 + d
    dow   = date.weekday()
    is_wk = dow >= 5
    is_fe = date in FERIADOS
    is_pk = dow in (4, 5)  # Vie, Sáb

    ws1.row_dimensions[row].height = 19

    shifts = {e: schedule[e].get(dow) for e in EMPLOYEES}
    active = sum(1 for s in shifts.values() if s)

    row_bg = C_FER if is_fe else (C_WKD if is_wk else C_WHI)

    # Fecha
    c = ws1.cell(row=row, column=1, value=date.strftime("%d/%m"))
    c.font = fn(bold=is_fe, sz=9, col=("A50000" if is_fe else ("444444" if is_wk else "000000")))
    c.fill = fl(row_bg)
    c.alignment = al()
    c.border = bdr()

    # Día
    c = ws1.cell(row=row, column=2, value=DAYS_ES[dow])
    c.font = fn(bold=True, sz=9, col=("A50000" if is_fe else ("595959" if is_wk else "000000")))
    c.fill = fl(row_bg)
    c.alignment = al()
    c.border = bdr()

    # Empleados
    for ei, emp in enumerate(EMPLOYEES, 3):
        sh = shifts[emp]
        c  = ws1.cell(row=row, column=ei)
        if sh is None:
            c.value = "FRANCO"
            c.font  = fn(sz=8, col="909090", it=True)
            c.fill  = fl(C_OFF)
        else:
            c.value = sh
            c.font  = fn(bold=True, sz=9, col=("A50000" if is_fe else "000000"))
            c.fill  = shift_fill(sh)
        c.alignment = al()
        c.border    = bdr()

    # Activos
    c = ws1.cell(row=row, column=8, value=active)
    c.font      = fn(bold=True, sz=9)
    c.fill      = fl(row_bg)
    c.alignment = al()
    c.border    = bdr()

    # Obs
    notes = []
    if is_fe: notes.append(f"⚑ {FERIADOS[date]}")
    if is_pk: notes.append("★ Día pico — req. 3 en guardia 12-22")
    c = ws1.cell(row=row, column=9, value="   ".join(notes))
    c.font      = fn(sz=8, it=True, col=("A50000" if is_fe else "595959"))
    c.fill      = fl(row_bg)
    c.alignment = al(h="left")
    c.border    = bdr()

# Leyenda
lr = 5 + 31 + 2
ws1.row_dimensions[lr].height = 5
ws1.row_dimensions[lr+1].height = 16
ws1.merge_cells(f"A{lr+1}:B{lr+1}")
c = ws1[f"A{lr+1}"]
c.value = "LEYENDA DE TURNOS"
c.font  = fn(bold=True, sz=8, col=C_HDR)

leg_items = [
    ("Turno Mañana", "9:00-17:00 / 9:00-18:00 / 9:00-19:00 / 10:00-18:00", C_MAN),
    ("Turno Centro", "12:00-21:00  /  12:00-22:00",                          C_TC),
    ("Turno Tarde",  "13:00-22:00",                                           C_TARD),
    ("Turno Noche",  "14:00-22:00",                                           C_NOCH),
    ("FRANCO",       "Día libre / descanso semanal",                          C_OFF),
]
for i, (lbl, desc, col) in enumerate(leg_items):
    r = lr + 2 + i
    ws1.row_dimensions[r].height = 14
    for ci, val in enumerate([lbl, desc], 1):
        c = ws1.cell(row=r, column=ci, value=val)
        c.font = fn(bold=(ci==1), sz=8)
        c.fill = fl(col)
        c.alignment = al(h=("center" if ci==1 else "left"))
        c.border = bdr()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — RESUMEN
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resumen")
ws2.sheet_view.showGridLines = False

for i, w in enumerate([16,10,12,14,18,26], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.row_dimensions[1].height = 28
ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value = "RESUMEN — JULIO 2026"
c.font  = fn(bold=True, sz=13, col=C_HDR)
c.fill  = fl(C_TIT)
c.alignment = al()

ws2.row_dimensions[3].height = 28
for col, h in enumerate(["Colaborador","Días\ntrabajados","Hs presencia\n(julio)","Hs semanales\n(contrato)","Francos semanales","Turno / horario habitual"], 1):
    hdr_cell(ws2, 3, col, h, wrap=True)

for ri, emp in enumerate(EMPLOYEES, 4):
    ws2.row_dimensions[ri].height = 18
    days_w = tot_h = 0
    for d in range(31):
        sh = schedule[emp].get((START + datetime.timedelta(days=d)).weekday())
        if sh:
            days_w += 1
            tot_h  += HRS.get(sh, 9)
    alt = "F2F9FF" if ri % 2 == 0 else C_WHI
    for ci, val in enumerate([emp, days_w, tot_h, f"{CONTRACT[emp]} hs", FRANCOS[emp], TURNOS[emp]], 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.font      = fn(bold=(ci==1), sz=9)
        c.fill      = fl(alt)
        c.alignment = al(h=("left" if ci in (1,5,6) else "center"))
        c.border    = bdr()

# Plantilla semanal
ws2.row_dimensions[10].height = 8
ws2.row_dimensions[11].height = 20
ws2.merge_cells("A11:G11")
c = ws2["A11"]
c.value = "PLANTILLA SEMANAL (se repite cada semana)"
c.font  = fn(bold=True, sz=10, col=C_HDR)

ws2.row_dimensions[12].height = 20
hdr_cell(ws2, 12, 1, "Colaborador", sz=9)
for ci, d in enumerate(DAYS_ES, 2):
    hdr_cell(ws2, 12, ci, d, sz=9)

for ri, emp in enumerate(EMPLOYEES, 13):
    ws2.row_dimensions[ri].height = 18
    c = ws2.cell(row=ri, column=1, value=emp)
    c.font = fn(bold=True, sz=9); c.alignment = al(h="left"); c.border = bdr()
    for ci, dow in enumerate(range(7), 2):
        sh = schedule[emp].get(dow)
        c  = ws2.cell(row=ri, column=ci)
        if sh is None:
            c.value = "FRANCO"; c.font = fn(sz=8, col="909090", it=True); c.fill = fl(C_OFF)
        else:
            c.value = sh; c.font = fn(bold=True, sz=9); c.fill = shift_fill(sh)
        c.alignment = al(); c.border = bdr()

# Cobertura por franja
ws2.row_dimensions[20].height = 8
ws2.row_dimensions[21].height = 20
ws2.merge_cells("A21:F21")
c = ws2["A21"]
c.value = "COBERTURA POR FRANJA HORARIA (días tipo)"
c.font  = fn(bold=True, sz=10, col=C_HDR)

ws2.row_dimensions[22].height = 18
for ci, h in enumerate(["Día","Franja","Cobertura","Personal presente"], 1):
    hdr_cell(ws2, 22, ci, h, bg=C_HDR2, sz=8)

cov_rows = [
    ("Lun / Mar",  "9-12",  "1 persona",    "Martina"),
    ("Lun / Mar",  "12-13", "2 personas",   "Martina + Yazmin"),
    ("Lun / Mar",  "13-17", "3 personas",   "Martina + Yazmin + Franco"),
    ("Lun / Mar",  "17-21", "2 personas",   "Yazmin + Franco"),
    ("Lun / Mar",  "21-22", "1 persona",    "Franco"),
    ("Mié",        "9-12",  "1 persona",    "Martina"),
    ("Mié",        "12-13", "2 personas",   "Martina + Yazmin"),
    ("Mié",        "13-17", "3 personas",   "Martina + Yazmin + Ana"),
    ("Mié",        "17-21", "2 personas",   "Yazmin + Ana"),
    ("Mié",        "21-22", "1 persona",    "Ana"),
    ("Jue",        "9-12",  "2 personas",   "Yazmin + Martina"),
    ("Jue",        "13-21", "2-4 personas", "Yazmin + Martina + Ana + Victoria"),
    ("Jue",        "21-22", "2 personas",   "Ana + Victoria"),
    ("Vie ★",      "9-12",  "2 personas",   "Franco + Yazmin"),
    ("Vie ★",      "12-21", "3-4 personas", "Franco + Yazmin + Ana (+ Victoria desde 14)"),
    ("Vie ★",      "21-22", "1 persona",    "Victoria"),
    ("Sáb ★",      "9-12",  "1 persona",    "Franco"),
    ("Sáb ★",      "12-14", "2 personas",   "Franco + Ana"),
    ("Sáb ★",      "14-19", "3 personas",   "Franco + Ana + Victoria  ← pico cubierto"),
    ("Sáb ★",      "19-22", "2 personas",   "Ana + Victoria"),
    ("Dom",        "10-12", "1 persona",    "Victoria"),
    ("Dom",        "12-13", "2 personas",   "Victoria + Franco"),
    ("Dom",        "13-18", "3 personas",   "Victoria + Franco + Ana"),
    ("Dom",        "18-21", "2 personas",   "Franco + Ana"),
    ("Dom",        "21-22", "2 personas",   "Franco + Ana"),
]

for ri, (dia, franja, cob, pers) in enumerate(cov_rows, 23):
    ws2.row_dimensions[ri].height = 14
    is_pk = "★" in dia
    bg = "FFF3E0" if is_pk else (C_WHI if ri % 2 else "F9F9F9")
    fc = "7B3000" if is_pk else "000000"
    for ci, val in enumerate([dia, franja, cob, pers], 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.font = fn(sz=8, bold=is_pk, col=fc)
        c.fill = fl(bg)
        c.alignment = al(h=("left" if ci == 4 else "center"))
        c.border = bdr()

# Notas al pie
nr = 23 + len(cov_rows) + 1
ws2.row_dimensions[nr].height = 14
ws2.merge_cells(f"A{nr}:F{nr}")
c = ws2[f"A{nr}"]
c.value = ("⚠  NOTA SÁBADOS: La cobertura simultánea de 3 personas se logra en la franja 14-19 hs. "
           "Si se requiere estrictamente 3 personas desde las 12 hs, se recomienda sumar colaborador o rotar a Yazmin los sábados.")
c.font = fn(sz=8, it=True, col="7B3000")
c.fill = fl("FFF3E0")
c.alignment = al(h="left", wrap=True)
c.border = bdr()
ws2.row_dimensions[nr].height = 28

out = "/mnt/user-data/outputs/guardia_julio_2026.xlsx"
wb.save(out)
print("OK →", out)
